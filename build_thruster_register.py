import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter

BV   = "003366"; BV2="336699"; AMBER="F4B942"; AMBERF="FFF3D6"
GREEN="C8E6C9"; GREENF="EAF7EA"; GREY="EEF2F7"; RED="CC0000"
wb = openpyxl.Workbook()

thin = Side(style="thin", color="C9D3DF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill("solid", fgColor=BV)
hdr_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(color=BV, bold=True, size=16)
sub_font = Font(color="555555", size=10, italic=True)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

# ---------------- Lists sheet (dropdown sources) ----------------
lst = wb.active; lst.title = "Lists"
lists = {
 "A":["Thruster type","Tunnel (transverse)","Retractable tunnel","Azimuth (rotating)","Retractable azimuth",
      "Azimuth CP","Podded","Water-jet","Voith-Schneider","Rim-driven","Pump-jet","Gill/hydro-jet","Steering grid/nozzle","Other"],
 "B":["Prime mover","Electric","Diesel","Hydraulic"],
 "C":["Discipline","General/Admin","Structural","Mechanical","Hydraulic","Electrical","Control","Torsional Vibration (TVC)","Testing","Other"],
 "D":["Status","Open - awaiting drawings","In review","Complete - all received","On hold","Closed"],
 "E":["Yes/No/NA","Yes","No","N/A"],
}
for col,vals in lists.items():
    for i,v in enumerate(vals, start=1):
        lst[f"{col}{i}"] = v
        if i==1: lst[f"{col}{i}"].font = Font(bold=True, color=BV)
lst.sheet_state = "hidden"

def dv(colletter, first, last=1000):
    # returns DataValidation referencing Lists column, skipping header row1
    end = {"Thruster type":14,"Prime mover":4,"Discipline":10,"Status":6,"Yes/No/NA":4}
    return None

def make_dv(list_col, n):
    d = DataValidation(type="list", formula1=f"=Lists!${list_col}$2:${list_col}${n}", allow_blank=True)
    return d

# ================= SOP sheet =================
sop = wb.create_sheet("Read me first (SOP)")
sop.sheet_view.showGridLines = False
sop.column_dimensions["A"].width = 3
sop.column_dimensions["B"].width = 110
def w(row, text, font=None, fill=None, height=None):
    c = sop.cell(row=row, column=2, value=text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if height: sop.row_dimensions[row].height = height
    return c

r=1
w(r,"Thruster Package Register — How to handle incoming submissions", title_font, height=24); r+=1
w(r,"Non-technical intake guide for the admin / shared-mailbox team. Keep this workbook on the shared drive so everyone updates the same file.", sub_font, height=28); r+=2

w(r,"THE GOLDEN RULE", Font(bold=True,color="FFFFFF",size=12), fill=BV, height=20); r+=1
w(r,"One thruster = ONE BV project registration = ONE review package — no matter how many separate deliveries (tranches) of drawings arrive over the months. Everything for that thruster is keyed to the CLIENT PO NUMBER (plus vessel/hull and thruster tag). You never need to understand the drawings — you only need to match the PO number.", Font(size=11), fill=GREY, height=58); r+=2

w(r,"WHY THIS MATTERS", Font(bold=True,color=BV,size=12), height=18); r+=1
w(r,"Thruster drawings do not all arrive at once. A typical thruster is delivered in tranches over ~6 months: Mechanical / structural first, then Hydraulic, then Electrical / Control, and the Torsional Vibration (TVC) usually LAST. If a later tranche is opened as a NEW package, one thruster ends up split across several packages — which is wrong. This workbook stops that.", Font(size=11), height=64); r+=2

w(r,"STEP-BY-STEP — every time a submission email arrives", Font(bold=True,color=BV,size=12), height=18); r+=1
steps = [
 "1.  Open the 'Intake Log' sheet. Add a new row.",
 "2.  Enter today's Date and the Client PO number (read it from the email / subject / cover letter).",
 "3.  Look at the 'ACTION' column — it fills in automatically:",
 "         ✓  EXISTING → route to BV-Reg-XXX   =  this PO is already registered. Do NOT create a new package.",
 "         ➤  NEW PACKAGE — create BV registration   =  this PO has never been seen before.",
 "4.  If EXISTING: forward the drawings to the review team under the BV Reg No shown, then go to 'Package Register',",
 "         find that row, and fill in the received-date for the discipline that just arrived (Mechanical / Hydraulic / etc.).",
 "5.  If NEW: create the BV project registration number as usual, add ONE row to 'Package Register',",
 "         then send the client the acknowledgement email (template below) so they quote the BV Reg No on all future tranches.",
 "6.  Never open a second registration for the same PO / vessel / thruster tag.",
]
for s in steps:
    w(r, s, Font(size=11, color=("15803D" if "EXISTING" in s else (RED if "NEW PACKAGE" in s else "1F2937"))), height=16); r+=1
r+=1

w(r,"IF THE CLIENT DID NOT QUOTE THE PO NUMBER", Font(bold=True,color=BV,size=12), height=18); r+=1
w(r,"Search the 'Package Register' (Ctrl+F) by Vessel / Hull number, Client name, or Thruster tag to find the existing registration. Still unsure which package it belongs to? Flag it to the responsible plan-approval surveyor before creating anything — do not guess by opening a new package.", Font(size=11), height=46); r+=2

w(r,"WHEN IS A PACKAGE COMPLETE?", Font(bold=True,color=BV,size=12), height=18); r+=1
w(r,"In 'Package Register', the 'Still outstanding' column lists the disciplines not yet received. When it is empty, all expected tranches (including the TVC) are in. The surveyor marks 'N/A' for any discipline a given thruster does not need (e.g. no hydraulics on a simple fixed-pitch electric tunnel thruster) — use the interactive Quick-Reference tool to see which disciplines that thruster type requires.", Font(size=11), height=52); r+=2

w(r,"ACKNOWLEDGEMENT EMAIL TEMPLATE (send on first receipt of a NEW package)", Font(bold=True,color="FFFFFF",size=12), fill=BV2, height=20); r+=1
tmpl = ("Subject:  BV Project Registration [BV-Reg-No] — Thruster [tag], Vessel [name/hull], PO [PO no]\n\n"
 "Dear [Client],\n\n"
 "We confirm receipt of your submission under Purchase Order [PO no]. This thruster has been registered under\n"
 "Bureau Veritas Project Registration No. [BV-Reg-No].\n\n"
 "Please quote this BV Registration Number — together with the PO number and thruster tag [tag] — on ALL future\n"
 "submissions for this thruster, including the later hydraulic, electrical/control and torsional-vibration (TVC)\n"
 "drawings. This ensures every delivery is added to the same review package.\n\n"
 "Kind regards,\n[Name] — Bureau Veritas Marine & Offshore, Plan Approval")
w(r, tmpl, Font(size=10.5, name="Consolas"), fill=GREY, height=200); r+=2
w(r,"Tip: keeping the client disciplined about quoting the BV Reg No is the single most effective fix — it removes the guesswork entirely.", sub_font, height=28)

# ================= Package Register =================
reg = wb.create_sheet("Package Register")
reg_cols = ["BV Reg No","Client PO No","Client / Yard","Vessel / Hull No","Thruster tag(s) & qty",
    "Thruster type","Prime mover","Date opened","Mechanical rcvd","Hydraulic rcvd","Electrical rcvd",
    "Control rcvd","TVC rcvd","Testing/other rcvd","Still outstanding","Status","Last update","Notes"]
reg.append(reg_cols)
style_header(reg,1,len(reg_cols))
reg.freeze_panes = "A2"
widths=[15,14,20,18,20,18,12,12,13,13,13,12,10,15,26,24,13,30]
for i,wd in enumerate(widths, start=1): reg.column_dimensions[get_column_letter(i)].width = wd
# 'Still outstanding' formula (col O = 15) referencing I..N (9..14)
for row in range(2, 300):
    f=(f'=IF(COUNTA(A{row})=0,"",TRIM('
       f'IF(I{row}="","Mechanical ","")&IF(J{row}="","Hydraulic ","")&IF(K{row}="","Electrical ","")&'
       f'IF(L{row}="","Control ","")&IF(M{row}="","TVC ","")&IF(N{row}="","Testing ","")))')
    reg.cell(row=row, column=15, value=f).alignment = wrap
# sample example row
ex=["BV-THR-2026-001","PO-88123","Damen / Yard 512","MV Northern Star / H-512","AZ-1 & AZ-2 (2 off)",
    "Azimuth (rotating)","Diesel","2026-01-15","2026-01-15","2026-03-02","","","","",None,"Open - awaiting drawings","2026-03-02","EXAMPLE ROW — delete. Awaiting Electrical, Control, TVC."]
reg.append(ex)
for c in range(1,len(reg_cols)+1):
    reg.cell(row=2,column=c).border=border; reg.cell(row=2,column=c).alignment=wrap
    reg.cell(row=2,column=c).font=Font(italic=True,color="888888")
# data validation
dv_type=make_dv("A",14); dv_prime=make_dv("B",4); dv_status=make_dv("D",6)
reg.add_data_validation(dv_type); dv_type.add(f"F2:F300")
reg.add_data_validation(dv_prime); dv_prime.add(f"G2:G300")
reg.add_data_validation(dv_status); dv_status.add(f"P2:P300")
# conditional format: Still outstanding empty (and reg exists) -> green ; not empty -> amber
reg.conditional_formatting.add("O2:O300",
    FormulaRule(formula=['AND($A2<>"",$O2="")'], fill=PatternFill("solid",fgColor=GREENF), font=Font(color="15803D",bold=True)))
reg.conditional_formatting.add("O2:O300",
    FormulaRule(formula=['AND($A2<>"",$O2<>"")'], fill=PatternFill("solid",fgColor=AMBERF), font=Font(color="9A6700")))
# highlight duplicate PO numbers in col B
reg.conditional_formatting.add("B2:B300",
    FormulaRule(formula=['AND($B2<>"",COUNTIF($B$2:$B$300,$B2)>1)'], fill=PatternFill("solid",fgColor="FFD6D6"), font=Font(color=RED,bold=True)))

# ================= Intake Log =================
il = wb.create_sheet("Intake Log")
il_cols=["Date received","Client PO No","BV Reg No (auto)","ACTION (auto)","Vessel/Hull (auto)","Thruster tag (auto)",
    "Sender / email","Drawings received (titles / count)","Discipline","Tranche #","Routed to","Notes"]
il.append(il_cols)
style_header(il,1,len(il_cols))
il.freeze_panes="A2"
il_w=[13,14,15,34,20,20,24,34,20,9,16,26]
for i,wd in enumerate(il_w,start=1): il.column_dimensions[get_column_letter(i)].width=wd
for row in range(2,600):
    # C: existing BV Reg No via INDEX/MATCH on Package Register (PO col B -> Reg col A)
    il.cell(row=row,column=3,value=(f"=IF($B{row}=\"\",\"\",IFERROR(INDEX('Package Register'!$A:$A,"
        f"MATCH($B{row},'Package Register'!$B:$B,0)),\"\"))"))
    # D: action text
    il.cell(row=row,column=4,value=(f'=IF($B{row}="","",IF($C{row}="","➤ NEW PACKAGE — create BV registration",'
        f'"✓ EXISTING — route to "&$C{row}))')).alignment=wrap
    # E vessel, F thruster tag auto
    il.cell(row=row,column=5,value=(f"=IF($C{row}=\"\",\"\",IFERROR(INDEX('Package Register'!$D:$D,"
        f"MATCH($B{row},'Package Register'!$B:$B,0)),\"\"))"))
    il.cell(row=row,column=6,value=(f"=IF($C{row}=\"\",\"\",IFERROR(INDEX('Package Register'!$E:$E,"
        f"MATCH($B{row},'Package Register'!$B:$B,0)),\"\"))"))
# discipline dropdown
dv_disc=make_dv("C",10); il.add_data_validation(dv_disc); dv_disc.add("I2:I600")
# conditional format ACTION column
il.conditional_formatting.add("D2:D600",
    FormulaRule(formula=['LEFT($D2,1)="✓"'], fill=PatternFill("solid",fgColor=GREENF), font=Font(color="15803D",bold=True)))
il.conditional_formatting.add("D2:D600",
    FormulaRule(formula=['LEFT($D2,1)="➤"'], fill=PatternFill("solid",fgColor=AMBERF), font=Font(color="9A6700",bold=True)))
# sample rows
il.cell(row=2,column=1,value="2026-01-15"); il.cell(row=2,column=2,value="PO-88123")
il.cell(row=2,column=7,value="engineering@damen.example"); il.cell(row=2,column=8,value="GA + propeller + shaft (5 dwgs)")
il.cell(row=2,column=9,value="Mechanical"); il.cell(row=2,column=10,value=1); il.cell(row=2,column=11,value="Machinery reviewer")
il.cell(row=3,column=1,value="2026-03-02"); il.cell(row=3,column=2,value="PO-88123")
il.cell(row=3,column=7,value="engineering@damen.example"); il.cell(row=3,column=8,value="HPU + hydraulic schematic (3 dwgs)")
il.cell(row=3,column=9,value="Hydraulic"); il.cell(row=3,column=10,value=2); il.cell(row=3,column=11,value="Machinery reviewer")
for row in (2,3):
    for c in range(1,len(il_cols)+1):
        il.cell(row=row,column=c).border=border; il.cell(row=row,column=c).alignment=wrap

# order sheets: SOP, Register, Intake, Lists(hidden)
wb.move_sheet("Read me first (SOP)", -(wb.sheetnames.index("Read me first (SOP)")))
order=["Read me first (SOP)","Package Register","Intake Log","Lists"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = wb.sheetnames.index("Read me first (SOP)")

wb.save("Thruster-Package-Register.xlsx")
print("saved. sheets:", wb.sheetnames)
